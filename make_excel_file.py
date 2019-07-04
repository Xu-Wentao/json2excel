# -*- coding: UTF-8 -*-

import openpyxl


def make_excel_file(file_name, sheets: list):
    """
    :param file_name: excel name you wanna create
    :param sheets: for example
     [{'title':'sheet0',
       'headers':{'name':'姓名','age':'年龄'},
       'data_list':[
            {'name':'admin','age':12},
            {'name':'guest','age':12},
       ]}]
    """
    wb = openpyxl.Workbook()
    try:
        for index, sheet in enumerate(sheets):
            if index == 0:
                ws = wb.worksheets[0]
                ws.title = sheet['title']
            else:
                ws = wb.create_sheet(sheet['title'])
            headers = sheet['headers']
            for i, value in enumerate(headers.values()):
                ws.cell(row=1, column=i + 1, value=value)
            for i, data in enumerate(sheet['data_list']):
                for j, key in enumerate(headers.keys()):
                    ws.cell(row=i + 2, column=j + 1, value=data[key])
        wb.save(filename=file_name)
        return file_name
    except Exception as e:
        raise ValueError(e)
